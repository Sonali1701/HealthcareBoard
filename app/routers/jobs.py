"""Job posting search/CRUD + applications + saved jobs."""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, File, HTTPException, Query, Response, UploadFile, status
from pydantic import BaseModel
from sqlalchemy import Integer, and_, delete, func, or_, select

from ..config import settings
from ..deps import CurrentUser, DbSession
from ..models import (
    Application,
    ApplicationEvent,
    Employer,
    EmployerMember,
    JobPosting,
    JobStatus,
    Notification,
    Offer,
    Profile,
    SavedJob,
    User,
)
from ..models.enums import ApplicationStatus, JobType, NotificationType
from ..services.email import send_new_application
from ..schemas.common import Message, Page
from ..schemas.job import (
    ApplicationCreate,
    ApplicationEventOut,
    ApplicationOut,
    ApplicationStageUpdate,
    JobCreate,
    JobOut,
    JobUpdate,
    SavedJobOut,
)

router = APIRouter(prefix="/api/jobs", tags=["jobs"])


def _require_job_manager(db: DbSession, job: JobPosting, user: CurrentUser) -> Employer:
    employer = db.get(Employer, job.employer_id)
    if employer.owner_user_id == user.user_id or user.role.value == "admin":
        return employer
    member = db.scalar(
        select(EmployerMember).where(
            EmployerMember.employer_id == employer.employer_id,
            EmployerMember.user_id == user.user_id,
        )
    )
    if not member:
        raise HTTPException(status_code=403, detail="Cannot manage this job")
    return employer


def _current_profile(db: DbSession, user: CurrentUser) -> Profile:
    profile = db.scalar(select(Profile).where(Profile.user_id == user.user_id))
    if not profile:
        raise HTTPException(status_code=400, detail="Create a profile first")
    return profile


# --- Search & CRUD --------------------------------------------------------

@router.get("", response_model=Page[JobOut])
def search_jobs(
    db: DbSession,
    user: CurrentUser,
    q: Optional[str] = Query(None, description="Full-text search"),
    specialty: Optional[str] = None,
    profession_type: Optional[str] = None,
    job_type: Optional[str] = None,
    state_code: Optional[str] = None,
    city: Optional[str] = None,
    pay_min: Optional[float] = None,
    is_urgent: Optional[bool] = None,
    employer_id: Optional[str] = None,
    facility: Optional[str] = None,
    group_openings: bool = Query(
        False, description="Collapse identical roles at the same facility into "
                           "one row carrying an `openings` count"),
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
):
    stmt = select(JobPosting).where(JobPosting.status == JobStatus.active)
    if q:
        stmt = stmt.where(JobPosting.search_text.like(f"%{q.lower()}%"))
    if specialty:
        stmt = stmt.where(JobPosting.specialty == specialty)
    if profession_type:
        stmt = stmt.where(JobPosting.profession_type == profession_type)
    if job_type:
        stmt = stmt.where(JobPosting.job_type == job_type)
    if state_code:
        stmt = stmt.where(JobPosting.state_code == state_code.upper())
    if city:
        stmt = stmt.where(JobPosting.city.ilike(f"%{city}%"))
    if pay_min is not None:
        stmt = stmt.where(JobPosting.pay_rate_max >= pay_min)
    if is_urgent is not None:
        stmt = stmt.where(JobPosting.is_urgent.is_(is_urgent))
    if employer_id:
        stmt = stmt.where(JobPosting.employer_id == employer_id)
    if facility:
        stmt = stmt.where(JobPosting.facility == facility)

    if group_openings:
        return _grouped_page(db, stmt, limit, offset)

    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(
        stmt.order_by(JobPosting.is_featured.desc(), JobPosting.created_at.desc())
        .limit(limit).offset(offset)
    ).all()
    return Page(items=rows, total=total, limit=limit, offset=offset)


# Agencies file one requisition per seat, so a single role at one facility can
# appear 30 times with 30 distinct req codes. They are NOT duplicates and must
# not be merged in the database — but the board is unreadable without folding
# them into a single row that says how many seats are open.
_GROUP_KEY = (JobPosting.title, JobPosting.facility, JobPosting.city,
              JobPosting.state_code, JobPosting.pay_rate_max)


def _grouped_page(db, stmt, limit: int, offset: int) -> Page:
    base = stmt.subquery()
    groups = (
        select(*[getattr(base.c, c.key) for c in _GROUP_KEY],
               func.count().label("openings"),
               func.min(base.c.job_id).label("job_id"))
        .group_by(*[getattr(base.c, c.key) for c in _GROUP_KEY])
    ).subquery()

    total = db.scalar(select(func.count()).select_from(groups)) or 0
    rows = db.execute(
        select(groups.c.job_id, groups.c.openings)
        .order_by(groups.c.openings.desc(), groups.c.job_id)
        .limit(limit).offset(offset)
    ).all()
    openings = {jid: n for jid, n in rows}
    jobs = db.scalars(
        select(JobPosting).where(JobPosting.job_id.in_(list(openings)))
    ).all() if openings else []
    items = []
    for job in sorted(jobs, key=lambda j: -openings[j.job_id]):
        out = JobOut.model_validate(job)
        out.openings = openings[job.job_id]
        items.append(out)
    return Page(items=items, total=total, limit=limit, offset=offset)


@router.post("", response_model=JobOut, status_code=status.HTTP_201_CREATED)
def create_job(employer_id: str, body: JobCreate, user: CurrentUser, db: DbSession):
    employer = db.get(Employer, employer_id)
    if not employer:
        raise HTTPException(status_code=404, detail="Employer not found")
    if employer.owner_user_id != user.user_id and user.role.value != "admin":
        member = db.scalar(
            select(EmployerMember).where(
                EmployerMember.employer_id == employer_id,
                EmployerMember.user_id == user.user_id,
            )
        )
        if not member:
            raise HTTPException(status_code=403, detail="Cannot post for this employer")

    job = JobPosting(employer_id=employer_id, posted_by_user_id=user.user_id,
                     **body.model_dump())
    job.rebuild_search_text()
    db.add(job)
    db.commit()
    db.refresh(job)
    return job


# --- Bulk upload / delete-all (Job Orders) --------------------------------
# Columns the Excel/CSV importer understands (header row, case-insensitive,
# spaces or underscores). Only `title` is required; everything else is optional.
_BULK_COLUMNS = (
    "title", "specialty", "profession_type", "job_type", "shift_type",
    "pay_rate_min", "pay_rate_max", "pay_unit", "housing_stipend",
    "signing_bonus", "city", "state_code", "years_exp_min", "is_urgent",
    "start_date", "description",
)
_BULK_FLOATS = {"pay_rate_min", "pay_rate_max", "housing_stipend", "signing_bonus"}
_BULK_MAX_ROWS = 1000

_STATE_CODES = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT", "delaware": "DE",
    "florida": "FL", "georgia": "GA", "hawaii": "HI", "idaho": "ID",
    "illinois": "IL", "indiana": "IN", "iowa": "IA", "kansas": "KS",
    "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN", "mississippi": "MS",
    "missouri": "MO", "montana": "MT", "nebraska": "NE", "nevada": "NV",
    "new hampshire": "NH", "new jersey": "NJ", "new mexico": "NM", "new york": "NY",
    "north carolina": "NC", "north dakota": "ND", "ohio": "OH", "oklahoma": "OK",
    "oregon": "OR", "pennsylvania": "PA", "rhode island": "RI", "south carolina": "SC",
    "south dakota": "SD", "tennessee": "TN", "texas": "TX", "utah": "UT",
    "vermont": "VT", "virginia": "VA", "washington": "WA", "west virginia": "WV",
    "wisconsin": "WI", "wyoming": "WY", "district of columbia": "DC",
    "washington dc": "DC", "d.c.": "DC",
}


def _norm_state(value) -> Optional[str]:
    """Normalise a state cell to a 2-letter code: accepts 'TX', 'texas', 'Texas'."""
    s = str(value).strip()
    if not s:
        return None
    if len(s) == 2 and s.isalpha():
        return s.upper()
    return _STATE_CODES.get(s.lower())  # None if unrecognised full name


def _require_employer_manager(db: DbSession, employer_id: str, user: CurrentUser) -> Employer:
    employer = db.get(Employer, employer_id)
    if not employer:
        raise HTTPException(status_code=404, detail="Employer not found")
    if employer.owner_user_id != user.user_id and user.role.value != "admin":
        member = db.scalar(select(EmployerMember).where(
            EmployerMember.employer_id == employer_id,
            EmployerMember.user_id == user.user_id))
        if not member:
            raise HTTPException(status_code=403, detail="Cannot manage this employer")
    return employer


def _parse_upload(data: bytes, filename: str) -> list[tuple[int, dict]]:
    """Read an .xlsx or .csv into ``(spreadsheet_row_number, {column: value})``
    tuples keyed by the header row. Unknown columns and empty cells are dropped;
    fully-blank rows are skipped but the real row numbers are preserved so error
    messages point at the right line in the user's file."""
    name = (filename or "").lower()
    rows: list[list] = []
    if name.endswith(".csv"):
        import csv
        text = data.decode("utf-8-sig", "replace")
        rows = [r for r in csv.reader(text.splitlines())]
    else:  # treat everything else as an Excel workbook
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        for r in ws.iter_rows(values_only=True):
            rows.append(list(r))
    if not rows:
        return []
    headers = [str(h or "").strip().lower().replace(" ", "_") for h in rows[0]]
    out = []
    for n, raw in enumerate(rows[1:], start=2):  # start=2 → row 1 is the header
        row = {}
        for h, v in zip(headers, raw):
            if h in _BULK_COLUMNS and v not in (None, ""):
                row[h] = v
        if row:
            out.append((n, row))
    return out


def _row_to_job(row: dict) -> tuple[Optional[dict], Optional[str]]:
    """Validate one parsed row into JobPosting kwargs, or (None, error)."""
    title = str(row.get("title") or "").strip()
    if not title:
        return None, None  # blank row → silently skipped, not an error
    job: dict = {"title": title[:300]}

    for col in ("specialty", "profession_type", "shift_type", "pay_unit",
                "city", "description"):
        if row.get(col) not in (None, ""):
            job[col] = str(row[col]).strip()
    if row.get("state_code") not in (None, ""):
        code = _norm_state(row["state_code"])
        if code is None:
            return None, f"state_code '{row['state_code']}' isn't a US state or 2-letter code"
        job["state_code"] = code

    jt = str(row.get("job_type") or "").strip().lower().replace("-", "_").replace(" ", "_")
    if jt:
        valid = {t.value for t in JobType}
        if jt not in valid:
            return None, f"job_type '{row['job_type']}' must be one of {', '.join(sorted(valid))}"
        job["job_type"] = JobType(jt)

    for col in _BULK_FLOATS:
        if row.get(col) not in (None, ""):
            try:
                job[col] = float(str(row[col]).replace("$", "").replace(",", "").strip())
            except ValueError:
                return None, f"{col} '{row[col]}' is not a number"
    if row.get("years_exp_min") not in (None, ""):
        try:
            job["years_exp_min"] = int(float(row["years_exp_min"]))
        except ValueError:
            return None, f"years_exp_min '{row['years_exp_min']}' is not a whole number"
    if row.get("is_urgent") not in (None, ""):
        job["is_urgent"] = str(row["is_urgent"]).strip().lower() in {"true", "yes", "1", "y"}
    if row.get("start_date") not in (None, ""):
        sd = row["start_date"]
        if isinstance(sd, datetime):
            job["start_date"] = sd.date()
        elif isinstance(sd, date):
            job["start_date"] = sd
        else:
            try:
                job["start_date"] = datetime.fromisoformat(str(sd)[:10]).date()
            except ValueError:
                return None, f"start_date '{sd}' must be YYYY-MM-DD"
    return job, None


@router.post("/bulk")
async def bulk_create_jobs(employer_id: str, user: CurrentUser, db: DbSession,
                           file: UploadFile = File(...)) -> dict:
    """Create many job orders at once from an uploaded Excel (.xlsx) or CSV."""
    _require_employer_manager(db, employer_id, user)
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="The file is empty.")
    if len(data) > 5 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 5 MB).")
    try:
        rows = _parse_upload(data, file.filename or "")
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=422, detail=f"Could not read the file: {exc}")
    if not rows:
        raise HTTPException(status_code=422,
                            detail="No rows found. Make sure row 1 has column headers "
                                   "(at least 'title') and jobs start on row 2.")
    if len(rows) > _BULK_MAX_ROWS:
        raise HTTPException(status_code=413,
                            detail=f"Too many rows ({len(rows)}). Limit is {_BULK_MAX_ROWS} per upload.")

    created, skipped, errors = 0, 0, []
    for rownum, row in rows:
        kwargs, err = _row_to_job(row)
        if err:
            errors.append(f"Row {rownum}: {err}")
            continue
        if not kwargs:
            skipped += 1
            continue
        job = JobPosting(employer_id=employer_id, posted_by_user_id=user.user_id, **kwargs)
        job.rebuild_search_text()
        db.add(job)
        created += 1
    if created:
        db.commit()
    return {"created": created, "skipped": skipped,
            "failed": len(errors), "errors": errors[:25]}


@router.get("/template")
def jobs_template(user: CurrentUser) -> Response:
    """Download a ready-to-fill Excel template for the bulk uploader."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"
    ws.append(list(_BULK_COLUMNS))
    ws.append([
        "ICU Registered Nurse", "ICU", "Nursing", "travel", "nights",
        2400, 2800, "weekly", 1200, 1500, "Dallas", "TX", 2, "TRUE",
        "2026-09-15", "13-week ICU travel contract, nights, 36 hrs/wk.",
    ])
    for i in range(1, len(_BULK_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=healthboard-jobs-template.xlsx"},
    )


@router.delete("/all")
def delete_all_jobs(employer_id: str, user: CurrentUser, db: DbSession) -> dict:
    """Permanently delete every job order for an employer, plus dependent rows
    (applications, their events, saved jobs, offers). Manager-only. This cannot
    be undone — the UI confirms before calling it."""
    _require_employer_manager(db, employer_id, user)
    job_ids = select(JobPosting.job_id).where(JobPosting.employer_id == employer_id)
    n = db.scalar(select(func.count()).select_from(job_ids.subquery())) or 0
    if n:
        app_ids = select(Application.application_id).where(Application.job_id.in_(job_ids))
        db.execute(delete(ApplicationEvent).where(ApplicationEvent.application_id.in_(app_ids)))
        db.execute(delete(Application).where(Application.job_id.in_(job_ids)))
        db.execute(delete(SavedJob).where(SavedJob.job_id.in_(job_ids)))
        db.execute(delete(Offer).where(Offer.job_id.in_(job_ids)))
        db.execute(delete(JobPosting).where(JobPosting.employer_id == employer_id))
        db.commit()
    return {"deleted": n}


# --- Job AI: natural-language job search (job seekers) ---------------------

_VALID_STATE_CODES = set(_STATE_CODES.values()) | {
    "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID",
    "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS",
    "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "OH", "OK",
    "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV",
    "WI", "WY", "DC",
}
_JOB_SHIFTS = {"night": "nights", "nights": "nights", "day": "days", "days": "days",
               "evening": "evenings", "evenings": "evenings", "rotating": "rotating"}
_JOB_TYPES = {"travel": "travel", "staff": "staff", "permanent": "staff",
              "full-time": "staff", "fulltime": "staff", "per diem": "per_diem",
              "perdiem": "per_diem", "prn": "per_diem", "contract": "contract"}
_JOB_FILLER = {"find", "show", "me", "jobs", "job", "roles", "role", "looking",
               "for", "want", "need", "search", "a", "an", "the", "in", "near",
               "with", "and", "or", "of", "to", "at", "on", "any", "some", "please",
               "positions", "position", "opening", "openings", "work", "working",
               # generic profession words → not useful as required keywords
               "nurse", "nursing", "rn", "lpn", "clinician",
               # pay / distance words handled elsewhere or noise
               "over", "above", "least", "paying", "pay", "min", "minimum", "per",
               "week", "weekly", "wk", "hour", "hourly", "hr", "month", "year",
               "yr", "salary", "rate", "that", "who", "are", "is", "my",
               # refinement / conversational filler
               "only", "just", "also", "more", "prefer", "rather", "instead",
               "actually", "really", "maybe", "give", "around", "about"}

_JOB_COPILOT_SYSTEM = (
    "You turn a healthcare job seeker's plain-English search into structured "
    "job-board filters. Output ONLY a JSON object, nothing else."
)
_JOB_COPILOT_INSTR = (
    "Return JSON with exactly these keys (use null when not specified):\n"
    '{"q":null,"specialty":null,"profession_type":null,"job_type":null,'
    '"state_code":null,"city":null,"pay_min":null,"shift_type":null,"is_urgent":null}\n\n'
    "- specialty: the clinical specialty as a short phrase ('ICU','telemetry',"
    "'med surg','emergency','labor delivery'). Normalise informal terms: "
    "'medsurg'/'medsurgian'->'med surg'; 'tele'->'telemetry'; 'l&d'->'labor delivery'; "
    "'peds'->'pediatric'; 'er'/'ed'->'emergency'. null if none.\n"
    "- profession_type: 'Nursing','Physician','Allied','APP', or 'Others' when clearly "
    "implied ('nurse'/'RN'->'Nursing'), else null.\n"
    "- job_type: one of 'travel','staff','per_diem','contract' ONLY when the "
    "arrangement is explicitly named ('permanent'/'full-time'->'staff'; "
    "'prn'/'per diem'->'per_diem'). Do NOT infer it from pay phrasing like "
    "'per week' or 'weekly' — that describes pay, not job type. null if unstated.\n"
    "- state_code: 2-letter US state code the job is in ('in Texas'->'TX'). null if none.\n"
    "- city: the city name only, no state ('near Dallas'->'Dallas'). null if none.\n"
    "- pay_min: minimum pay as a plain number when a floor is given ('over $2500/week',"
    "'at least 2500','2500+','paying 2500'). Digits only, no symbols. null if none.\n"
    "- shift_type: 'days','nights','evenings', or 'rotating' when stated. null if none.\n"
    "- is_urgent: true only if they explicitly want urgent/ASAP roles, else null.\n"
    "- q: any remaining useful keywords not captured above (certifications, extra "
    "qualifiers). Lowercase, space-separated. null if none.\n"
    "IGNORE words like 'find','show me','jobs','roles','looking for','I want'."
)


# The LLM classifies a query's profession into a broad category; the imported
# job feed stores specific profession titles. Map category -> the title
# fragments that belong to it, so a "nursing" search actually hits RN/LPN jobs.
_PROFESSION_MATCH = {
    "nursing": ["RN", "LPN", "LVN", "Nurse"],
    "nurse": ["RN", "LPN", "LVN", "Nurse"],
    "physician": ["Physician", "MD", "DO"],
    "app": ["Nurse Practitioner", "CRNA", "Physician Assistant", "Anesthetist"],
    "allied": ["Physical Therapy", "Occupational Therapy", "Speech", "Radiology",
               "Imaging", "Respiratory", "Pharmacy", "Lab", "Medical Assistant",
               "Sonographer", "Tech", "Dietary", "Social"],
}


class JobCopilotQuery(BaseModel):
    message: str
    context: Optional[dict] = None


def _job_rule_filters(message: str) -> dict:
    """Lightweight fallback extractor so Job AI still works with the LLM off."""
    text = " " + message.lower() + " "
    out: dict = {}

    m = re.search(r"\$?\s*(\d{3,6})\s*(?:\+|/?\s*(?:wk|week|weekly|hr|hour|hourly|k))",
                  text)
    if not m:
        m = re.search(r"(?:over|above|at least|paying|pay|min(?:imum)?)\s*\$?\s*(\d{3,6})",
                      text)
    if m:
        val = int(m.group(1))
        out["pay_min"] = float(val * 1000 if val < 100 else val)  # "2k" style guard

    for word, canon in _JOB_SHIFTS.items():
        if re.search(rf"\b{word}\b", text):
            out["shift_type"] = canon
            break
    for word, canon in _JOB_TYPES.items():
        if word in text:
            out["job_type"] = canon
            break
    if re.search(r"\b(urgent|asap|immediate(?:ly)?)\b", text):
        out["is_urgent"] = True

    # State: full names first (may be two words), then bare 2-letter codes.
    for name, code in _STATE_CODES.items():
        if re.search(rf"\b{re.escape(name)}\b", text):
            out["state_code"] = code
            break
    if "state_code" not in out:
        for raw in re.findall(r"\b([a-z]{2})\b", message.lower()):
            if raw.upper() in _VALID_STATE_CODES and raw not in ("in", "or", "me", "hi", "so", "no"):
                out["state_code"] = raw.upper()
                break

    # Keywords: scrub out everything already captured as a structured filter so
    # words like 'texas', 'over', '$2500/week' or 'travel' never become required
    # search terms (which would AND-match to zero results).
    scrub = " " + message.lower() + " "
    scrub = re.sub(r"\$?\s*\d[\d,]*\s*(?:\+|/?\s*(?:wk|week|weekly|hr|hour|hourly|k|month|yr|year))?",
                   " ", scrub)
    for name in _STATE_CODES:
        scrub = re.sub(rf"\b{re.escape(name)}\b", " ", scrub)
    if out.get("state_code"):
        scrub = re.sub(rf"\b{out['state_code'].lower()}\b", " ", scrub)
    for w in list(_JOB_SHIFTS) + list(_JOB_TYPES) + ["urgent", "asap", "immediate", "immediately"]:
        scrub = re.sub(rf"\b{re.escape(w)}\b", " ", scrub)
    kw = [t for t in re.findall(r"[a-z][a-z&/-]+", scrub)
          if len(t) >= 2 and t not in _JOB_FILLER][:6]
    if kw:
        out["q"] = " ".join(kw)
    return out


def _job_copilot_filters(raw: dict) -> dict:
    """Validate/clean an LLM (or rule) filter dict into safe job filters."""
    out: dict = {}
    if not isinstance(raw, dict):
        return out

    def s(key, cap=80):
        v = raw.get(key)
        if v in (None, "", "null"):
            return None
        return str(v).strip()[:cap]

    for key in ("q", "specialty", "profession_type", "city"):
        v = s(key)
        if v:
            out[key] = v
    sc = s("state_code", 20)
    if sc:
        code = _norm_state(sc)
        if code:
            out["state_code"] = code
    jt = (s("job_type", 20) or "").lower().replace("-", "_").replace(" ", "_")
    if jt in {t.value for t in JobType}:
        out["job_type"] = jt
    sh = (s("shift_type", 20) or "").lower()
    for word, canon in _JOB_SHIFTS.items():
        if word in sh:
            out["shift_type"] = canon
            break
    pay = raw.get("pay_min")
    if pay not in (None, "", "null"):
        try:
            out["pay_min"] = float(re.sub(r"[^\d.]", "", str(pay)) or 0) or None
            if out["pay_min"] is None:
                out.pop("pay_min")
        except ValueError:
            pass
    if str(raw.get("is_urgent")).lower() in ("true", "1", "yes"):
        out["is_urgent"] = True
    return out


_JOB_MERGE_FIELDS = ("q", "specialty", "profession_type", "job_type",
                     "state_code", "city", "pay_min", "shift_type", "is_urgent")
_JOB_RESET_RE = re.compile(r"\b(start over|reset|new search|clear|never mind)\b", re.I)


def _merge_job_filters(prior: dict, delta: dict) -> dict:
    out = dict(prior)
    for k in _JOB_MERGE_FIELDS:
        if k in delta:
            if delta[k] in (None, ""):
                out.pop(k, None)
            else:
                out[k] = delta[k]
    return out


def _job_copilot_summary(f: dict, total: int) -> str:
    bits = []
    if f.get("specialty"):
        bits.append(f["specialty"])
    elif f.get("q"):
        bits.append(f"“{f['q']}”")
    if f.get("profession_type") and not f.get("specialty"):
        bits.append(f["profession_type"])
    if f.get("job_type"):
        bits.append(str(f["job_type"]).replace("_", " "))
    place = ", ".join(x for x in ((f.get("city") or "").title() or None,
                                  f.get("state_code")) if x)
    if place:
        bits.append(place)
    if f.get("pay_min"):
        bits.append(f"${int(f['pay_min']):,}+")
    if f.get("shift_type"):
        bits.append(str(f["shift_type"]))
    if f.get("is_urgent"):
        bits.append("urgent")
    label = " · ".join(bits)
    if not label:
        return ("Tell me what you're looking for — e.g. “ICU night travel jobs in "
                "Texas paying over $2,500/week”.")
    if total:
        return f"Found {total:,} job{'' if total == 1 else 's'} matching {label}."
    return (f"No jobs matched {label} right now. Try widening the pay, area, "
            "or shift.")


@router.post("/copilot")
def job_copilot(body: JobCopilotQuery, user: CurrentUser, db: DbSession):
    """Natural-language job search for job seekers (the 'Job AI' assistant)."""
    message = (body.message or "").strip()
    if not message:
        return {"answer": _job_copilot_summary({}, 0), "filters": {}, "items": [], "total": 0}

    rule_delta = _job_rule_filters(message)
    llm_delta: dict = {}
    if settings.llm_enabled and settings.llm_api_key and settings.llm_model:
        from ..clean_names_llm import _llm
        try:
            raw = _llm(message, system=_JOB_COPILOT_SYSTEM, instr=_JOB_COPILOT_INSTR,
                       max_chars=500, retries=1, timeout=8)
            llm_delta = _job_copilot_filters(raw or {})
        except Exception:  # noqa: BLE001
            llm_delta = {}
    delta = {**rule_delta, **llm_delta}
    if llm_delta:
        delta["q"] = llm_delta.get("q")  # LLM is authority on free-text keywords

    prior = {} if _JOB_RESET_RE.search(message) else _job_copilot_filters(body.context or {})
    filters = _merge_job_filters(prior, delta)

    stmt = select(JobPosting).where(JobPosting.status == JobStatus.active)
    tokens = []
    if filters.get("specialty"):
        tokens += filters["specialty"].lower().split()
    if filters.get("q"):
        tokens += filters["q"].lower().split()
    for tok in [t for t in tokens if len(t) >= 2][:6]:
        stmt = stmt.where(JobPosting.search_text.like(f"%{tok}%"))
    if filters.get("state_code"):
        stmt = stmt.where(JobPosting.state_code == filters["state_code"])
    if filters.get("city"):
        stmt = stmt.where(JobPosting.city.ilike(f"%{filters['city']}%"))
    # Profession: the LLM emits a broad category ("Nursing"), but the feed uses
    # specific titles ("RN", "LPN/LVN", "Physician"…). Map the category to those
    # so it narrows correctly instead of matching nothing. Unknown categories are
    # ignored (better a broad result than zero), and NULL professions are kept.
    prof_pats = _PROFESSION_MATCH.get(str(filters.get("profession_type") or "").lower())
    if prof_pats:
        stmt = stmt.where(or_(
            JobPosting.profession_type.is_(None),
            *[JobPosting.profession_type.ilike(f"%{p}%") for p in prof_pats]))
    # Pay and shift are SOFT preferences: most of the feed lists no pay and some
    # no shift, so a seeker who mentions either must not be shown zero jobs. Only
    # exclude a job that EXPLICITLY falls short; a job with no value still shows.
    if filters.get("pay_min"):
        stmt = stmt.where(or_(JobPosting.pay_rate_max.is_(None),
                              JobPosting.pay_rate_max >= filters["pay_min"]))
    if filters.get("job_type"):
        stmt = stmt.where(JobPosting.job_type == JobType(filters["job_type"]))
    if filters.get("shift_type"):
        sh = str(filters["shift_type"]).rstrip("s")
        stmt = stmt.where(or_(JobPosting.shift_type.is_(None),
                              JobPosting.shift_type.ilike(f"%{sh}%")))
    if filters.get("is_urgent"):
        stmt = stmt.where(JobPosting.is_urgent.is_(True))

    page = _grouped_page(db, stmt, limit=12, offset=0)
    return {"answer": _job_copilot_summary(filters, page.total),
            "filters": filters, "items": page.items, "total": page.total}


@router.get("/recommended", response_model=Page[JobOut])
def recommended_jobs(user: CurrentUser, db: DbSession,
                     limit: int = Query(20, ge=1, le=50)):
    """Roles that fit the signed-in professional's own profile.

    The platform already scores candidates against a job for recruiters; this
    runs the same comparison from the other side. It is deliberately ordered
    rather than filtered — a nurse with an unusual specialty should still see
    the closest roles instead of an empty page.
    """
    from ..models import Profile

    profile = db.scalar(select(Profile).where(Profile.user_id == user.user_id))
    stmt = select(JobPosting).where(JobPosting.status == JobStatus.active)

    if not profile or not (profile.profession_type or profile.specialty
                           or profile.state_code):
        # Nothing to match on yet — show the newest roles rather than nothing,
        # and let the profile-completion prompt do the rest.
        rows = db.scalars(stmt.order_by(JobPosting.created_at.desc()).limit(limit)).all()
        return Page(items=rows, total=len(rows), limit=limit, offset=0)

    # Rank by how much of the profile a role matches, strongest signal first:
    # the licence has to line up for the job to be doable at all, specialty
    # decides whether they are competitive, and location decides whether they
    # would take it.
    score = (
        func.coalesce(
            (JobPosting.profession_type == profile.profession_type).cast(Integer) * 4, 0)
        + func.coalesce(
            (JobPosting.specialty == profile.specialty).cast(Integer) * 3, 0)
        + func.coalesce(
            (JobPosting.state_code == profile.state_code).cast(Integer) * 2, 0)
        + func.coalesce(
            (func.lower(JobPosting.city) == func.lower(profile.city)).cast(Integer), 0)
    ).label("fit")

    rows = db.execute(
        select(JobPosting, score)
        .where(JobPosting.status == JobStatus.active)
        .order_by(score.desc(), JobPosting.is_featured.desc(),
                  JobPosting.created_at.desc())
        .limit(limit)
    ).all()
    items = []
    for job, fit in rows:
        out = JobOut.model_validate(job)
        out.fit_score = int(fit or 0)
        items.append(out)
    return Page(items=items, total=len(items), limit=limit, offset=0)


@router.get("/{job_id}", response_model=JobOut)
def get_job(job_id: str, db: DbSession, user: CurrentUser):
    job = db.get(JobPosting, job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    job.view_count += 1
    db.commit()
    db.refresh(job)
    return job


@router.get("/{job_id}/pay-estimate")
async def job_pay_estimate(job_id: str, db: DbSession, user: CurrentUser):
    """Estimated weekly take-home for this job, GSA per-diem-aware.

    Treats the job's advertised hourly as a blended package (the clinician's own
    pay, not a bill rate) and splits it into taxable pay + tax-free per-diem for
    the job's location — the same model as the seeker Pay Tools tab. California
    daily overtime applies automatically.
    """
    from ..schemas.gsa import PayPackageRequest
    from ..services.gsa import calculate_pay_package, get_gsa_rates

    job = db.get(JobPosting, job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    reqs = job.requirements or {}
    hours = float(reqs.get("hours_per_week") or 36) or 36
    hourly = float(job.pay_rate_max or job.pay_rate_min or 0)
    if not hourly and reqs.get("weekly_pay"):
        try:
            hourly = float(reqs["weekly_pay"]) / hours
        except (TypeError, ValueError, ZeroDivisionError):
            hourly = 0.0
    if not hourly or not job.state_code:
        return {"available": False}
    weeks = int(reqs.get("duration_weeks") or 13) or 13
    try:
        payreq = PayPackageRequest(
            bill_rate=hourly, city=(job.city or job.state_code), state_code=job.state_code,
            margin_pct=0, burden_multiplier=1.0, hours_per_week=hours,
            contract_weeks=min(max(weeks, 1), 104), tax_rate=0.22)
    except Exception:  # noqa: BLE001 — a bad rate/location just means "no estimate"
        return {"available": False}
    rates = await get_gsa_rates(payreq.city, payreq.state_code)
    r = calculate_pay_package(payreq, rates)
    pd = r.option_perdiem
    return {
        "available": True,
        "hourly": round(hourly, 2), "hours_per_week": hours,
        "weekly_net": pd.est_weekly_net, "weekly_tax_free": pd.weekly_tax_free,
        "weekly_taxable_gross": pd.weekly_taxable_gross, "weekly_total": pd.weekly_total,
        "overtime": r.breakdown["overtime"], "gsa_live": rates.source == "api.gsa.gov",
        "city": job.city, "state_code": job.state_code,
    }


@router.patch("/{job_id}", response_model=JobOut)
def update_job(job_id: str, body: JobUpdate, user: CurrentUser, db: DbSession):
    job = db.get(JobPosting, job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    _require_job_manager(db, job, user)
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(job, field, value)
    job.rebuild_search_text()
    db.commit()
    db.refresh(job)
    return job


@router.delete("/{job_id}", status_code=204)
def delete_job(job_id: str, user: CurrentUser, db: DbSession):
    job = db.get(JobPosting, job_id)
    if not job:
        return
    _require_job_manager(db, job, user)
    job.status = JobStatus.closed
    db.commit()


# --- Applications ---------------------------------------------------------

@router.post("/{job_id}/apply", response_model=ApplicationOut, status_code=201)
def apply_to_job(job_id: str, body: ApplicationCreate, user: CurrentUser, db: DbSession):
    job = db.get(JobPosting, job_id)
    if not job or job.status != JobStatus.active:
        raise HTTPException(status_code=404, detail="Job not available")
    profile = (
        db.get(Profile, body.profile_id) if body.profile_id
        else _current_profile(db, user)
    )
    if not profile:
        raise HTTPException(status_code=400, detail="Profile not found")

    existing = db.scalar(
        select(Application).where(
            and_(Application.job_id == job_id, Application.profile_id == profile.profile_id)
        )
    )
    if existing:
        raise HTTPException(status_code=409, detail="Already applied to this job")

    app = Application(
        job_id=job_id,
        profile_id=profile.profile_id,
        cover_letter=body.cover_letter,
        resume_snapshot_url=body.resume_snapshot_url or profile.resume_url,
        source=body.source,
    )
    db.add(app)
    job.application_count += 1
    db.flush()
    db.add(ApplicationEvent(application_id=app.application_id,
                            to_status=ApplicationStatus.applied.value,
                            actor_user_id=user.user_id))
    # Notify the employer owner, in-app and by email.
    employer = db.get(Employer, job.employer_id)
    candidate_name = f"{profile.first_name} {profile.last_name}".strip()
    if employer:
        db.add(Notification(
            user_id=employer.owner_user_id,
            type=NotificationType.application,
            title="New application",
            body=f"{candidate_name} applied to {job.title}",
            data={"job_id": job_id, "application_id": app.application_id},
        ))
    db.commit()
    db.refresh(app)
    if employer:
        owner = db.get(User, employer.owner_user_id)
        if owner and owner.email:
            send_new_application(owner.email, candidate_name, job.title)
    return app


@router.get("/{job_id}/applications", response_model=list[ApplicationOut])
def list_job_applications(job_id: str, user: CurrentUser, db: DbSession,
                          status_filter: Optional[ApplicationStatus] = Query(None, alias="status")):
    job = db.get(JobPosting, job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    _require_job_manager(db, job, user)
    stmt = select(Application).where(Application.job_id == job_id)
    if status_filter:
        stmt = stmt.where(Application.status == status_filter)
    return db.scalars(stmt.order_by(Application.applied_at.desc())).all()


# The candidate ATS pipeline, in the order applicants actually move through it.
# "rejected"/"withdrawn" are terminal and sit outside the ordered steps.
_APPLICANT_STAGES = ["applied", "screening", "interview", "offer", "hired"]


@router.get("/{job_id}/applicants")
def list_job_applicants(job_id: str, user: CurrentUser, db: DbSession,
                        status_filter: Optional[ApplicationStatus] = Query(None, alias="status")):
    """Applicants for a job, with each candidate's details attached.

    `/applications` returns bare rows carrying a profile_id, which is unusable
    as a review screen. Applying to a job reveals the candidate to that
    employer (the apply notification already names them), so this returns their
    real name and contact to the job's recruiters — the counterpart to the
    credit-gated reveal that governs the cold directory.
    """
    job = db.get(JobPosting, job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    _require_job_manager(db, job, user)

    stmt = select(Application).where(Application.job_id == job_id)
    if status_filter:
        stmt = stmt.where(Application.status == status_filter)
    apps = db.scalars(stmt.order_by(Application.applied_at.desc())).all()

    profs = {p.profile_id: p for p in db.scalars(
        select(Profile).where(Profile.profile_id.in_([a.profile_id for a in apps])))} if apps else {}

    items, by_status = [], {}
    for a in apps:
        p = profs.get(a.profile_id)
        status = a.status.value if hasattr(a.status, "value") else str(a.status)
        by_status[status] = by_status.get(status, 0) + 1
        items.append({
            "application_id": a.application_id,
            "profile_id": a.profile_id,
            "user_id": p.user_id if p else None,
            "name": f"{p.first_name} {p.last_name}".strip() if p else "Candidate",
            "headline": p.headline if p else None,
            "profession_type": p.profession_type if p else None,
            "specialty": p.specialty if p else None,
            "years_experience": p.years_experience if p else None,
            "location": ", ".join(x for x in [getattr(p, "city", None),
                                              getattr(p, "state_code", None)] if x) if p else None,
            "completion": p.completion_score if p else 0,
            "email": p.email if p else None,
            "phone": p.phone if p else None,
            "resume_url": p.resume_url if p else None,
            "cover_letter": a.cover_letter,
            "recruiter_rating": a.recruiter_rating,
            "recruiter_notes": a.recruiter_notes,
            "status": status,
            "stage_index": _APPLICANT_STAGES.index(status) if status in _APPLICANT_STAGES else None,
            "stages": _APPLICANT_STAGES,
            "is_closed": status in {"rejected", "withdrawn"},
            "applied_at": a.applied_at,
            "status_updated_at": a.status_updated_at,
        })
    return {
        "job": {"job_id": job.job_id, "title": job.title,
                "specialty": job.specialty, "profession_type": job.profession_type,
                "location": ", ".join(x for x in [job.city, job.state_code] if x)},
        "items": items,
        "by_status": by_status,
        "stages": _APPLICANT_STAGES,
    }


# --- Saved jobs -----------------------------------------------------------

@router.post("/{job_id}/save", response_model=SavedJobOut, status_code=201)
def save_job(job_id: str, user: CurrentUser, db: DbSession):
    job = db.get(JobPosting, job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    profile = _current_profile(db, user)
    existing = db.scalar(
        select(SavedJob).where(
            and_(SavedJob.job_id == job_id, SavedJob.profile_id == profile.profile_id)
        )
    )
    if existing:
        return existing
    saved = SavedJob(job_id=job_id, profile_id=profile.profile_id)
    db.add(saved)
    db.commit()
    db.refresh(saved)
    return saved


@router.delete("/{job_id}/save", status_code=204)
def unsave_job(job_id: str, user: CurrentUser, db: DbSession):
    profile = _current_profile(db, user)
    saved = db.scalar(
        select(SavedJob).where(
            and_(SavedJob.job_id == job_id, SavedJob.profile_id == profile.profile_id)
        )
    )
    if saved:
        db.delete(saved)
        db.commit()
